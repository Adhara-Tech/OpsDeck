#!/usr/bin/env python3
"""
BCDR Export - OpsDeck Disaster Recovery
=======================================
Standalone script to extract all data from the OpsDeck database
and generate an independent Excel file (.xlsx) per entity type.

Generates a directory with ~56 Excel files covering:
  - Users, groups, locations
  - Assets, peripherals, software, licenses
  - Subscriptions, suppliers, contacts, contracts, purchases, budgets
  - Business services and components
  - Risks, incidents, changes
  - Frameworks, controls, compliance evidence
  - Audits, policies
  - BCDR plans, credentials, certificates
  - Links, documentation, tags
  - Onboarding/offboarding, security activities

Usage:
    python scripts/bcdr_export.py --database-url postgresql://opsdeck:opsdeck@localhost:5432/opsdeck
    python scripts/bcdr_export.py --host localhost --port 5432 --db opsdeck
    python scripts/bcdr_export.py --output /backups/opsdeck_20260224

Requirements:
    pip install sqlalchemy psycopg2-binary openpyxl
"""

import argparse
import os
import sys
from datetime import datetime, date, time
from pathlib import Path

try:
    from sqlalchemy import create_engine, text
except ImportError:
    print("ERROR: sqlalchemy not installed. Run: pip install sqlalchemy")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ─── Excel Styles ────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

INFO_FONT = Font(name="Calibri", italic=True, size=9, color="666666")
INFO_ALIGNMENT = Alignment(horizontal="center")

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")
DATA_ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

EMPTY_FONT = Font(name="Calibri", italic=True, color="999999", size=10)
EMPTY_ALIGNMENT = Alignment(horizontal="center")

# Status colors (green = OK, yellow = in progress, red = problem)
STATUS_COLORS = {
    # Green - OK/completed
    "active": "C6EFCE", "operational": "C6EFCE", "compliant": "C6EFCE",
    "passed": "C6EFCE", "resolved": "C6EFCE", "closed": "C6EFCE",
    "mitigated": "C6EFCE", "completed": "C6EFCE", "pass": "C6EFCE",
    "approved": "C6EFCE", "success": "C6EFCE", "verified": "C6EFCE",
    "in use": "C6EFCE", "in_use": "C6EFCE", "available": "C6EFCE",
    "won": "C6EFCE", "locked": "C6EFCE", "sent": "C6EFCE",
    # Yellow - in progress / warning
    "pending": "FFEB9C", "in progress": "FFEB9C", "in_progress": "FFEB9C",
    "investigating": "FFEB9C", "in treatment": "FFEB9C", "provisioning": "FFEB9C",
    "evaluating": "FFEB9C", "draft": "FFEB9C", "identified": "FFEB9C",
    "assessed": "FFEB9C", "contained": "FFEB9C", "observation": "FFEB9C",
    "planned": "FFEB9C", "prep": "FFEB9C", "auditor review": "FFEB9C",
    "in review": "FFEB9C", "negotiating": "FFEB9C", "poc": "FFEB9C",
    "scheduled": "FFEB9C", "pending review": "FFEB9C", "standard": "FFEB9C",
    # Red - problem / inactive
    "expired": "FFC7CE", "archived": "FFC7CE", "failed": "FFC7CE",
    "fail": "FFC7CE", "gap": "FFC7CE", "rejected": "FFC7CE",
    "retired": "FFC7CE", "issue_detected": "FFC7CE", "flagged": "FFC7CE",
    "lost": "FFC7CE", "cancelled": "FFC7CE", "qualified": "FFC7CE",
    # Severity-specific
    "sev-0": "FFC7CE", "sev-1": "FFC7CE",
    "sev-2": "FFEB9C", "sev-3": "D6E4F0",
    "emergency": "FFC7CE", "normal": "FFEB9C",
}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def fmt_date(d):
    """Format date/datetime to DD/MM/YYYY."""
    if d is None:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d)


def fmt_datetime(d):
    """Format datetime to DD/MM/YYYY HH:MM."""
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%d/%m/%Y %H:%M")
    if isinstance(d, date):
        return d.strftime("%d/%m/%Y")
    return str(d)


def fmt_bool(v):
    """Format boolean to Yes/No."""
    if v is None:
        return ""
    return "Yes" if v else "No"


def fmt_val(v):
    """Format a generic value for Excel."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return fmt_bool(v)
    if isinstance(v, datetime):
        return fmt_datetime(v)
    if isinstance(v, date):
        return fmt_date(v)
    if isinstance(v, time):
        return v.strftime("%H:%M")
    return v


def apply_header_style(ws, row, num_cols):
    """Apply header style to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_data_style(cell, center=True):
    """Apply data style to a cell."""
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGNMENT if center else DATA_ALIGNMENT_LEFT
    cell.border = THIN_BORDER


def apply_status_color(cell, value):
    """Apply conditional color based on status value."""
    if value is None:
        return
    key = str(value).strip().lower()
    color = STATUS_COLORS.get(key)
    if color:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def auto_width(ws):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = min(max(max_length + 3, 10), 50)


# ─── Sheet Definitions ──────────────────────────────────────────────────────
# Each sheet is a dict with:
#   name        - tab name (max 31 chars)
#   tab_color   - hex color for the tab
#   query       - SQL query
#   headers     - list of column headers
#   status_cols - 0-based indices of status columns (for coloring)
#   text_cols   - 0-based indices of long text columns (left-aligned)

def get_sheet_configs():
    """Return the configuration for all export sheets."""
    return [
        # ── Organization ─────────────────────────────────────────────────
        {
            "name": "Users",
            "tab_color": "2F5496",
            "query": """
                SELECT u.id, u.name, u.email, u.role, u.department, u.job_title,
                       u.personal_email,
                       CASE WHEN u.is_archived THEN 'Archived' ELSE 'Active' END as status,
                       u.created_at,
                       m.name as manager, b.name as buddy
                FROM "user" u
                LEFT JOIN "user" m ON u.manager_id = m.id
                LEFT JOIN "user" b ON u.buddy_id = b.id
                ORDER BY u.name
            """,
            "headers": ["ID", "Name", "Email", "Role", "Department", "Job Title",
                        "Personal Email", "Status", "Created",
                        "Manager", "Buddy"],
            "status_cols": [7],
            "text_cols": [],
        },
        {
            "name": "Groups",
            "tab_color": "2F5496",
            "query": """
                SELECT g.id, g.name, g.description,
                       COUNT(ug.user_id) as member_count,
                       STRING_AGG(u.name, ', ' ORDER BY u.name) as members
                FROM "group" g
                LEFT JOIN user_groups ug ON g.id = ug.group_id
                LEFT JOIN "user" u ON ug.user_id = u.id
                GROUP BY g.id, g.name, g.description
                ORDER BY g.name
            """,
            "headers": ["ID", "Name", "Description", "Members Count", "Members"],
            "status_cols": [],
            "text_cols": [2, 4],
        },
        {
            "name": "Locations",
            "tab_color": "2F5496",
            "query": """
                SELECT id, name, address, city, zip_code, country, timezone,
                       phone, reception_email,
                       CASE WHEN is_archived THEN 'Archived' ELSE 'Active' END as status,
                       created_at
                FROM location
                ORDER BY name
            """,
            "headers": ["ID", "Name", "Address", "City", "Zip Code", "Country",
                        "Timezone", "Phone", "Reception Email", "Status", "Created"],
            "status_cols": [9],
            "text_cols": [2],
        },

        # ── Assets ───────────────────────────────────────────────────────
        {
            "name": "Assets",
            "tab_color": "548235",
            "query": """
                SELECT a.id, a.internal_id, a.name, br.name as brand, am.name as model,
                       a.serial_number, a.status,
                       CASE WHEN a.is_critical THEN 'Yes' ELSE 'No' END as critical,
                       CASE WHEN a.is_virtual THEN 'Yes' ELSE 'No' END as virtual,
                       a.purchase_date, a.cost, a.currency, a.warranty_length,
                       u.name as assigned_user, l.name as location,
                       s.name as supplier, a.comments,
                       CASE WHEN a.is_archived THEN 'Archived' ELSE 'Active' END as archive_status,
                       a.created_at
                FROM asset a
                LEFT JOIN "user" u ON a.user_id = u.id
                LEFT JOIN location l ON a.location_id = l.id
                LEFT JOIN supplier s ON a.supplier_id = s.id
                LEFT JOIN brand br ON a.brand_id = br.id
                LEFT JOIN asset_model am ON a.model_id = am.id
                ORDER BY a.name
            """,
            "headers": ["ID", "Internal ID", "Name", "Brand", "Model",
                        "Serial Number", "Status", "Critical", "Virtual",
                        "Purchase Date", "Cost", "Currency", "Warranty (months)",
                        "Assigned User", "Location", "Supplier", "Comments",
                        "Archive Status", "Created"],
            "status_cols": [6, 17],
            "text_cols": [16],
        },
        {
            "name": "Peripherals",
            "tab_color": "548235",
            "query": """
                SELECT p.id, p.name, p.type, br.name as brand, p.serial_number,
                       p.status, p.purchase_date, p.cost, p.currency,
                       p.warranty_length,
                       u.name as assigned_user, a.name as parent_asset,
                       l.name as location, s.name as supplier,
                       CASE WHEN p.is_archived THEN 'Archived' ELSE 'Active' END as archive_status,
                       p.created_at
                FROM peripheral p
                LEFT JOIN "user" u ON p.user_id = u.id
                LEFT JOIN asset a ON p.asset_id = a.id
                LEFT JOIN location l ON p.location_id = l.id
                LEFT JOIN supplier s ON p.supplier_id = s.id
                LEFT JOIN brand br ON p.brand_id = br.id
                ORDER BY p.name
            """,
            "headers": ["ID", "Name", "Type", "Brand", "Serial Number",
                        "Status", "Purchase Date", "Cost", "Currency",
                        "Warranty (months)", "Assigned User", "Parent Asset",
                        "Location", "Supplier", "Archive Status", "Created"],
            "status_cols": [5, 14],
            "text_cols": [],
        },
        {
            "name": "Software",
            "tab_color": "548235",
            "query": """
                SELECT sw.id, sw.name, sw.category, sw.description,
                       s.name as supplier,
                       CASE WHEN sw.is_archived THEN 'Archived' ELSE 'Active' END as status,
                       sw.created_at
                FROM software sw
                LEFT JOIN supplier s ON sw.supplier_id = s.id
                ORDER BY sw.name
            """,
            "headers": ["ID", "Name", "Category", "Description", "Supplier",
                        "Status", "Created"],
            "status_cols": [5],
            "text_cols": [3],
        },
        {
            "name": "Licenses",
            "tab_color": "548235",
            "query": """
                SELECT li.id, li.name, li.license_key, li.cost, li.currency,
                       li.purchase_date, li.expiry_date,
                       u.name as assigned_user, sw.name as software,
                       sub.name as subscription,
                       CASE WHEN li.is_archived THEN 'Archived' ELSE 'Active' END as status,
                       li.created_at
                FROM license li
                LEFT JOIN "user" u ON li.user_id = u.id
                LEFT JOIN software sw ON li.software_id = sw.id
                LEFT JOIN subscription sub ON li.subscription_id = sub.id
                ORDER BY li.name
            """,
            "headers": ["ID", "Name", "License Key", "Cost", "Currency",
                        "Purchase Date", "Expiry Date", "Assigned User",
                        "Software", "Subscription", "Status", "Created"],
            "status_cols": [10],
            "text_cols": [2],
        },
        {
            "name": "Maintenance Logs",
            "tab_color": "548235",
            "query": """
                SELECT ml.id, ml.event_type, ml.description, ml.status,
                       ml.event_date, ml.ticket_link, ml.notes,
                       u.name as assigned_to,
                       a.name as asset, p.name as peripheral,
                       ml.created_at
                FROM maintenance_log ml
                LEFT JOIN "user" u ON ml.assigned_to_id = u.id
                LEFT JOIN asset a ON ml.asset_id = a.id
                LEFT JOIN peripheral p ON ml.peripheral_id = p.id
                ORDER BY ml.event_date DESC
            """,
            "headers": ["ID", "Event Type", "Description", "Status",
                        "Event Date", "Ticket Link", "Notes", "Assigned To",
                        "Asset", "Peripheral", "Created"],
            "status_cols": [3],
            "text_cols": [2, 6],
        },
        {
            "name": "Disposal Records",
            "tab_color": "548235",
            "query": """
                SELECT dr.id, dr.disposal_date, dr.disposal_method,
                       dr.disposal_partner, dr.notes,
                       a.name as asset, p.name as peripheral
                FROM disposal_record dr
                LEFT JOIN asset a ON dr.asset_id = a.id
                LEFT JOIN peripheral p ON dr.peripheral_id = p.id
                ORDER BY dr.disposal_date DESC
            """,
            "headers": ["ID", "Disposal Date", "Method", "Partner", "Notes",
                        "Asset", "Peripheral"],
            "status_cols": [],
            "text_cols": [4],
        },

        # ── Procurement & Vendors ────────────────────────────────────────
        {
            "name": "Suppliers",
            "tab_color": "BF8F00",
            "query": """
                SELECT id, name, email, phone, address, website,
                       compliance_status, gdpr_dpa_signed,
                       security_assessment_completed, data_storage_region,
                       compliance_notes,
                       CASE WHEN is_archived THEN 'Archived' ELSE 'Active' END as status,
                       created_at
                FROM supplier
                ORDER BY name
            """,
            "headers": ["ID", "Name", "Email", "Phone", "Address", "Website",
                        "Compliance Status", "GDPR DPA Signed",
                        "Security Assessment Done", "Data Storage Region",
                        "Compliance Notes", "Status", "Created"],
            "status_cols": [6, 11],
            "text_cols": [4, 10],
        },
        {
            "name": "Contacts",
            "tab_color": "BF8F00",
            "query": """
                SELECT c.id, c.name, c.email, c.phone, c.role,
                       s.name as supplier,
                       CASE WHEN c.is_archived THEN 'Archived' ELSE 'Active' END as status,
                       c.created_at
                FROM contact c
                LEFT JOIN supplier s ON c.supplier_id = s.id
                ORDER BY c.name
            """,
            "headers": ["ID", "Name", "Email", "Phone", "Role", "Supplier",
                        "Status", "Created"],
            "status_cols": [6],
            "text_cols": [],
        },
        {
            "name": "Subscriptions",
            "tab_color": "BF8F00",
            "query": """
                SELECT sub.id, sub.name, sub.subscription_type, sub.description,
                       sub.cost, sub.currency, sub.pricing_model,
                       sub.cost_per_user,
                       sub.renewal_date, sub.renewal_period_type,
                       sub.auto_renew,
                       s.name as supplier, sw.name as software,
                       u.name as owner, b.name as budget,
                       CASE WHEN sub.is_archived THEN 'Archived' ELSE 'Active' END as status,
                       sub.created_at
                FROM subscription sub
                LEFT JOIN supplier s ON sub.supplier_id = s.id
                LEFT JOIN software sw ON sub.software_id = sw.id
                LEFT JOIN "user" u ON sub.user_id = u.id
                LEFT JOIN budget b ON sub.budget_id = b.id
                ORDER BY sub.name
            """,
            "headers": ["ID", "Name", "Type", "Description", "Cost", "Currency",
                        "Pricing Model", "Cost/User", "Renewal Date",
                        "Renewal Period", "Auto Renew", "Supplier", "Software",
                        "Owner", "Budget", "Status", "Created"],
            "status_cols": [15],
            "text_cols": [3],
        },
        {
            "name": "Contracts",
            "tab_color": "BF8F00",
            "query": """
                SELECT co.id, co.name, co.status, co.contract_type,
                       s.name as supplier, co.contact_email,
                       co.cost, co.currency, co.payment_frequency,
                       co.start_date, co.end_date,
                       co.notice_period_days, co.is_auto_renew,
                       co.renewal_notes, co.description, co.created_at
                FROM contract co
                LEFT JOIN supplier s ON co.supplier_id = s.id
                ORDER BY co.name
            """,
            "headers": ["ID", "Name", "Status", "Type", "Supplier",
                        "Contact Email", "Cost", "Currency", "Payment Freq.",
                        "Start Date", "End Date", "Notice Period (days)",
                        "Auto Renew", "Renewal Notes", "Description", "Created"],
            "status_cols": [2],
            "text_cols": [13, 14],
        },
        {
            "name": "Purchases",
            "tab_color": "BF8F00",
            "query": """
                SELECT p.id, p.internal_id, p.description, p.invoice_number,
                       p.purchase_date, p.comments,
                       s.name as supplier, pm.name as payment_method,
                       b.name as budget,
                       CASE WHEN p.is_archived THEN 'Archived' ELSE 'Active' END as status,
                       p.created_at
                FROM purchase p
                LEFT JOIN supplier s ON p.supplier_id = s.id
                LEFT JOIN payment_method pm ON p.payment_method_id = pm.id
                LEFT JOIN budget b ON p.budget_id = b.id
                ORDER BY p.purchase_date DESC
            """,
            "headers": ["ID", "Internal ID", "Description", "Invoice Number",
                        "Purchase Date", "Comments", "Supplier", "Payment Method",
                        "Budget", "Status", "Created"],
            "status_cols": [9],
            "text_cols": [2, 5],
        },
        {
            "name": "Budgets",
            "tab_color": "BF8F00",
            "query": """
                SELECT id, name, category, amount, currency,
                       period, valid_from, valid_until,
                       CASE WHEN is_archived THEN 'Archived' ELSE 'Active' END as status,
                       created_at
                FROM budget
                ORDER BY name
            """,
            "headers": ["ID", "Name", "Category", "Amount", "Currency",
                        "Period", "Valid From", "Valid Until", "Status", "Created"],
            "status_cols": [8],
            "text_cols": [],
        },
        {
            "name": "Payment Methods",
            "tab_color": "BF8F00",
            "query": """
                SELECT pm.id, pm.name, pm.method_type, pm.details,
                       pm.expiry_date, u.name as owner,
                       CASE WHEN pm.is_archived THEN 'Archived' ELSE 'Active' END as status,
                       pm.created_at
                FROM payment_method pm
                LEFT JOIN "user" u ON pm.user_id = u.id
                ORDER BY pm.name
            """,
            "headers": ["ID", "Name", "Type", "Details", "Expiry Date",
                        "Owner", "Status", "Created"],
            "status_cols": [6],
            "text_cols": [3],
        },

        # ── Business Services ────────────────────────────────────────────
        {
            "name": "Business Services",
            "tab_color": "7030A0",
            "query": """
                SELECT bs.id, bs.name, bs.description, bs.category,
                       bs.criticality, bs.status,
                       u.name as owner,
                       cc.name as cost_center,
                       bs.sla_response_hours, bs.sla_resolution_hours,
                       bs.created_at
                FROM business_service bs
                LEFT JOIN "user" u ON bs.owner_id = u.id
                LEFT JOIN cost_center cc ON bs.cost_center_id = cc.id
                ORDER BY bs.criticality, bs.name
            """,
            "headers": ["ID", "Name", "Description", "Category", "Criticality",
                        "Status", "Owner", "Cost Center",
                        "SLA Response (h)", "SLA Resolution (h)", "Created"],
            "status_cols": [5],
            "text_cols": [2],
        },
        {
            "name": "Service Components",
            "tab_color": "7030A0",
            "query": """
                SELECT sc.id, bs.name as service, sc.component_type,
                       sc.component_id, sc.notes, sc.created_at
                FROM service_component sc
                LEFT JOIN business_service bs ON sc.service_id = bs.id
                ORDER BY bs.name, sc.component_type
            """,
            "headers": ["ID", "Service", "Component Type", "Component ID",
                        "Notes", "Created"],
            "status_cols": [],
            "text_cols": [4],
        },
        {
            "name": "Configurations",
            "tab_color": "7030A0",
            "query": """
                SELECT c.id, c.name, c.description,
                       c.owner_type, c.owner_id,
                       bs.name as service, sw.name as software,
                       a.name as asset
                FROM configuration c
                LEFT JOIN business_service bs ON c.service_id = bs.id
                LEFT JOIN software sw ON c.software_id = sw.id
                LEFT JOIN asset a ON c.asset_id = a.id
                ORDER BY c.name
            """,
            "headers": ["ID", "Name", "Description", "Owner Type", "Owner ID",
                        "Service", "Software", "Asset"],
            "status_cols": [],
            "text_cols": [2],
        },

        # ── Risk Management ──────────────────────────────────────────────
        {
            "name": "Risks",
            "tab_color": "C00000",
            "query": """
                SELECT r.id, r.risk_description, r.extended_description,
                       tt.name as threat_type, r.status,
                       r.treatment_strategy, u.name as owner,
                       r.inherent_impact, r.inherent_likelihood,
                       r.residual_impact, r.residual_likelihood,
                       r.mitigation_plan, r.next_review_date,
                       r.link, r.created_at
                FROM risk r
                LEFT JOIN threat_type tt ON r.threat_type_id = tt.id
                LEFT JOIN "user" u ON r.owner_id = u.id
                ORDER BY r.status, r.id
            """,
            "headers": ["ID", "Description", "Extended Description",
                        "Threat Type", "Status", "Treatment Strategy", "Owner",
                        "Inherent Impact", "Inherent Likelihood",
                        "Residual Impact", "Residual Likelihood",
                        "Mitigation Plan", "Next Review", "Link", "Created"],
            "status_cols": [4],
            "text_cols": [1, 2, 11],
        },
        {
            "name": "Threat Types",
            "tab_color": "C00000",
            "query": """
                SELECT id, name, category, description
                FROM threat_type
                ORDER BY category, name
            """,
            "headers": ["ID", "Name", "Category", "Description"],
            "status_cols": [],
            "text_cols": [3],
        },

        # ── Incidents & Changes ──────────────────────────────────────────
        {
            "name": "Incidents",
            "tab_color": "FF6600",
            "query": """
                SELECT si.id, si.external_ref, si.title, si.description,
                       si.incident_date, si.status, si.severity, si.impact,
                       si.data_breach, si.third_party_impacted,
                       rep.name as reported_by, own.name as owner,
                       asg.name as assignee,
                       si.resolved_at, si.created_at
                FROM security_incident si
                LEFT JOIN "user" rep ON si.reported_by_id = rep.id
                LEFT JOIN "user" own ON si.owner_id = own.id
                LEFT JOIN "user" asg ON si.assignee_id = asg.id
                ORDER BY si.incident_date DESC
            """,
            "headers": ["ID", "External Ref", "Title", "Description",
                        "Incident Date", "Status", "Severity", "Impact",
                        "Data Breach", "Third Party Impacted",
                        "Reported By", "Owner", "Assignee",
                        "Resolved At", "Created"],
            "status_cols": [5, 6],
            "text_cols": [3],
        },
        {
            "name": "Post-Incident Reviews",
            "tab_color": "FF6600",
            "query": """
                SELECT pir.id, si.external_ref as incident_ref, si.title as incident,
                       pir.summary, pir.lead_up, pir.fault,
                       pir.impact_analysis, pir.detection, pir.response,
                       pir.recovery, pir.lessons_learned,
                       pir.is_locked, pir.created_at
                FROM post_incident_review pir
                LEFT JOIN security_incident si ON pir.incident_id = si.id
                ORDER BY pir.created_at DESC
            """,
            "headers": ["ID", "Incident Ref", "Incident Title",
                        "Summary", "Lead Up", "Fault",
                        "Impact Analysis", "Detection", "Response",
                        "Recovery", "Lessons Learned", "Locked", "Created"],
            "status_cols": [],
            "text_cols": [3, 4, 5, 6, 7, 8, 9, 10],
        },
        {
            "name": "Changes",
            "tab_color": "FF6600",
            "query": """
                SELECT ch.id, ch.external_ref, ch.title, ch.change_type,
                       ch.priority, ch.risk_impact, ch.status,
                       ch.requires_approval, ch.description,
                       ch.implementation_plan, ch.rollback_plan,
                       ch.scheduled_start, ch.scheduled_end,
                       req.name as requester, asg.name as assignee,
                       appr.name as approved_by, ch.approved_at,
                       bs.name as service,
                       ch.created_at, ch.executed_at, ch.closed_at
                FROM change ch
                LEFT JOIN "user" req ON ch.requester_id = req.id
                LEFT JOIN "user" asg ON ch.assignee_id = asg.id
                LEFT JOIN "user" appr ON ch.approved_by_id = appr.id
                LEFT JOIN business_service bs ON ch.service_id = bs.id
                ORDER BY ch.created_at DESC
            """,
            "headers": ["ID", "External Ref", "Title", "Type", "Priority",
                        "Risk Impact", "Status", "Requires Approval",
                        "Description", "Implementation Plan", "Rollback Plan",
                        "Scheduled Start", "Scheduled End",
                        "Requester", "Assignee", "Approved By", "Approved At",
                        "Service", "Created", "Executed At", "Closed At"],
            "status_cols": [6],
            "text_cols": [8, 9, 10],
        },

        # ── Compliance ───────────────────────────────────────────────────
        {
            "name": "Frameworks",
            "tab_color": "4472C4",
            "query": """
                SELECT id, name, description, link,
                       is_custom, is_active
                FROM framework
                ORDER BY name
            """,
            "headers": ["ID", "Name", "Description", "Link", "Custom", "Active"],
            "status_cols": [],
            "text_cols": [2],
        },
        {
            "name": "Framework Controls",
            "tab_color": "4472C4",
            "query": """
                SELECT fc.id, f.name as framework, fc.control_id,
                       fc.name, fc.description,
                       fc.is_applicable, fc.soa_justification
                FROM framework_control fc
                LEFT JOIN framework f ON fc.framework_id = f.id
                ORDER BY f.name, fc.control_id
            """,
            "headers": ["ID", "Framework", "Control ID", "Name", "Description",
                        "Applicable", "SOA Justification"],
            "status_cols": [],
            "text_cols": [4, 6],
        },
        {
            "name": "Compliance Links",
            "tab_color": "4472C4",
            "query": """
                SELECT cl.id, f.name as framework,
                       fc.control_id, fc.name as control_name,
                       cl.linkable_type, cl.linkable_id, cl.description
                FROM compliance_link cl
                LEFT JOIN framework_control fc ON cl.framework_control_id = fc.id
                LEFT JOIN framework f ON fc.framework_id = f.id
                ORDER BY f.name, fc.control_id
            """,
            "headers": ["ID", "Framework", "Control ID", "Control Name",
                        "Linked Type", "Linked ID", "Description"],
            "status_cols": [],
            "text_cols": [6],
        },
        {
            "name": "Compliance Rules",
            "tab_color": "4472C4",
            "query": """
                SELECT cr.id, f.name as framework,
                       fc.control_id, fc.name as control_name,
                       cr.name, cr.description,
                       cr.target_model, cr.criteria,
                       cr.frequency_days, cr.grace_period_days,
                       cr.enabled, cr.created_at
                FROM compliance_rule cr
                LEFT JOIN framework_control fc ON cr.framework_control_id = fc.id
                LEFT JOIN framework f ON fc.framework_id = f.id
                ORDER BY f.name, fc.control_id
            """,
            "headers": ["ID", "Framework", "Control ID", "Control Name",
                        "Rule Name", "Description", "Target Model", "Criteria",
                        "Frequency (days)", "Grace Period (days)",
                        "Enabled", "Created"],
            "status_cols": [],
            "text_cols": [5, 7],
        },

        # ── Audits ───────────────────────────────────────────────────────
        {
            "name": "Audits",
            "tab_color": "4472C4",
            "query": """
                SELECT ca.id, ca.name, ca.status, ca.outcome,
                       ca.audit_type, ca.start_date, ca.end_date,
                       f.name as framework,
                       co.name as auditor,
                       u.name as internal_lead,
                       ca.locked_at, ca.created_at
                FROM compliance_audit ca
                LEFT JOIN framework f ON ca.framework_id = f.id
                LEFT JOIN contact co ON ca.auditor_id = co.id
                LEFT JOIN "user" u ON ca.internal_lead_id = u.id
                ORDER BY ca.start_date DESC
            """,
            "headers": ["ID", "Name", "Status", "Outcome", "Type",
                        "Start Date", "End Date", "Framework",
                        "Auditor", "Internal Lead", "Locked At", "Created"],
            "status_cols": [2, 3],
            "text_cols": [],
        },
        {
            "name": "Audit Items",
            "tab_color": "4472C4",
            "query": """
                SELECT aci.id, ca.name as audit,
                       aci.control_code, aci.control_title,
                       aci.control_description,
                       aci.is_applicable, aci.justification,
                       aci.status,
                       aci.internal_comments, aci.auditor_findings
                FROM audit_control_item aci
                LEFT JOIN compliance_audit ca ON aci.audit_id = ca.id
                ORDER BY ca.name, aci.control_code
            """,
            "headers": ["ID", "Audit", "Control Code", "Control Title",
                        "Description", "Applicable", "Justification",
                        "Status", "Internal Comments", "Auditor Findings"],
            "status_cols": [7],
            "text_cols": [4, 6, 8, 9],
        },

        # ── Policies ─────────────────────────────────────────────────────
        {
            "name": "Policies",
            "tab_color": "4472C4",
            "query": """
                SELECT id, title, category, description, link, created_at
                FROM policy
                ORDER BY title
            """,
            "headers": ["ID", "Title", "Category", "Description", "Link", "Created"],
            "status_cols": [],
            "text_cols": [3],
        },
        {
            "name": "Policy Versions",
            "tab_color": "4472C4",
            "query": """
                SELECT pv.id, p.title as policy, pv.version_number,
                       pv.status, pv.effective_date, pv.end_date,
                       pv.content
                FROM policy_version pv
                LEFT JOIN policy p ON pv.policy_id = p.id
                ORDER BY p.title, pv.version_number DESC
            """,
            "headers": ["ID", "Policy", "Version", "Status",
                        "Effective Date", "End Date", "Content"],
            "status_cols": [3],
            "text_cols": [6],
        },

        # ── BCDR ─────────────────────────────────────────────────────────
        {
            "name": "BCDR Plans",
            "tab_color": "00B050",
            "query": """
                SELECT id, name, description, created_at
                FROM bcdr_plan
                ORDER BY name
            """,
            "headers": ["ID", "Name", "Description", "Created"],
            "status_cols": [],
            "text_cols": [2],
        },
        {
            "name": "BCDR Tests",
            "tab_color": "00B050",
            "query": """
                SELECT bt.id, bp.name as plan, bt.test_date,
                       bt.status, bt.notes,
                       u.name as assignee
                FROM bcdr_test_log bt
                LEFT JOIN bcdr_plan bp ON bt.plan_id = bp.id
                LEFT JOIN "user" u ON bt.assignee_id = u.id
                ORDER BY bt.test_date DESC
            """,
            "headers": ["ID", "Plan", "Test Date", "Status", "Notes", "Assignee"],
            "status_cols": [3],
            "text_cols": [4],
        },

        # ── Credentials & Certificates ───────────────────────────────────
        {
            "name": "Credentials",
            "tab_color": "7030A0",
            "query": """
                SELECT cr.id, cr.name, cr.type, cr.break_glass,
                       cr.description, cr.owner_type, cr.owner_id,
                       sw.name as software, a.name as asset,
                       cr.created_at, cr.updated_at
                FROM credentials cr
                LEFT JOIN software sw ON cr.software_id = sw.id
                LEFT JOIN asset a ON cr.asset_id = a.id
                ORDER BY cr.name
            """,
            "headers": ["ID", "Name", "Type", "Break Glass", "Description",
                        "Owner Type", "Owner ID", "Software", "Asset",
                        "Created", "Updated"],
            "status_cols": [],
            "text_cols": [4],
        },
        {
            "name": "Credential Secrets",
            "tab_color": "7030A0",
            "query": """
                SELECT cs.id, cr.name as credential,
                       cs.masked_value, cs.is_active,
                       cs.created_at, cs.expires_at
                FROM credential_secrets cs
                LEFT JOIN credentials cr ON cs.credential_id = cr.id
                ORDER BY cr.name, cs.created_at DESC
            """,
            "headers": ["ID", "Credential", "Masked Value", "Active",
                        "Created", "Expires At"],
            "status_cols": [],
            "text_cols": [],
        },
        {
            "name": "Certificates",
            "tab_color": "7030A0",
            "query": """
                SELECT c.id, c.name, c.type, c.description,
                       c.owner_type, c.owner_id,
                       c.created_at, c.updated_at
                FROM certificates c
                ORDER BY c.name
            """,
            "headers": ["ID", "Name", "Type", "Description",
                        "Owner Type", "Owner ID", "Created", "Updated"],
            "status_cols": [],
            "text_cols": [3],
        },
        {
            "name": "Certificate Versions",
            "tab_color": "7030A0",
            "query": """
                SELECT cv.id, c.name as certificate,
                       cv.version_notes, cv.valid_from, cv.expires_at,
                       cv.issuer, cv.common_name, cv.serial_number,
                       cv.private_key_location, cv.is_active, cv.created_at
                FROM certificate_versions cv
                LEFT JOIN certificates c ON cv.certificate_id = c.id
                ORDER BY c.name, cv.created_at DESC
            """,
            "headers": ["ID", "Certificate", "Notes", "Valid From", "Expires At",
                        "Issuer", "Common Name", "Serial Number",
                        "Private Key Location", "Active", "Created"],
            "status_cols": [],
            "text_cols": [2],
        },

        # ── Links & Documentation ────────────────────────────────────────
        {
            "name": "Links",
            "tab_color": "00B0F0",
            "query": """
                SELECT l.id, l.name, l.description, l.url,
                       l.owner_type, l.owner_id,
                       sw.name as software, l.created_at
                FROM link l
                LEFT JOIN software sw ON l.software_id = sw.id
                ORDER BY l.name
            """,
            "headers": ["ID", "Name", "Description", "URL",
                        "Owner Type", "Owner ID", "Software", "Created"],
            "status_cols": [],
            "text_cols": [2, 3],
        },
        {
            "name": "Documents",
            "tab_color": "00B0F0",
            "query": """
                SELECT d.id, d.name, d.description, d.external_link,
                       d.owner_type, d.owner_id,
                       sw.name as software, d.created_at
                FROM documentation d
                LEFT JOIN software sw ON d.software_id = sw.id
                ORDER BY d.name
            """,
            "headers": ["ID", "Name", "Description", "External Link",
                        "Owner Type", "Owner ID", "Software", "Created"],
            "status_cols": [],
            "text_cols": [2],
        },
        {
            "name": "Tags",
            "tab_color": "00B0F0",
            "query": """
                SELECT id, name,
                       CASE WHEN is_archived THEN 'Archived' ELSE 'Active' END as status
                FROM tag
                ORDER BY name
            """,
            "headers": ["ID", "Name", "Status"],
            "status_cols": [2],
            "text_cols": [],
        },
        {
            "name": "Attachments",
            "tab_color": "00B0F0",
            "query": """
                SELECT id, filename, secure_filename,
                       linkable_type, linkable_id, created_at
                FROM attachment
                ORDER BY created_at DESC
            """,
            "headers": ["ID", "Filename", "Secure Filename",
                        "Linked Type", "Linked ID", "Created"],
            "status_cols": [],
            "text_cols": [],
        },

        # ── Onboarding / Offboarding ────────────────────────────────────
        {
            "name": "Onboarding",
            "tab_color": "00B050",
            "query": """
                SELECT op.id, op.external_ref, op.new_hire_name,
                       u.name as user_account, op.target_email,
                       op.personal_email, op.start_date, op.status,
                       pk.name as pack,
                       m.name as manager, b.name as buddy,
                       op.created_at
                FROM onboarding_process op
                LEFT JOIN "user" u ON op.user_id = u.id
                LEFT JOIN onboarding_pack pk ON op.pack_id = pk.id
                LEFT JOIN "user" m ON op.assigned_manager_id = m.id
                LEFT JOIN "user" b ON op.assigned_buddy_id = b.id
                ORDER BY op.start_date DESC
            """,
            "headers": ["ID", "External Ref", "New Hire Name", "User Account",
                        "Target Email", "Personal Email", "Start Date", "Status",
                        "Pack", "Manager", "Buddy", "Created"],
            "status_cols": [7],
            "text_cols": [],
        },
        {
            "name": "Offboarding",
            "tab_color": "00B050",
            "query": """
                SELECT ofp.id, u.name as employee, m.name as manager,
                       ofp.departure_date, ofp.status, ofp.notes,
                       ofp.created_at
                FROM offboarding_process ofp
                LEFT JOIN "user" u ON ofp.user_id = u.id
                LEFT JOIN "user" m ON ofp.manager_id = m.id
                ORDER BY ofp.departure_date DESC
            """,
            "headers": ["ID", "Employee", "Manager", "Departure Date",
                        "Status", "Notes", "Created"],
            "status_cols": [4],
            "text_cols": [5],
        },

        # ── Security Activities ──────────────────────────────────────────
        {
            "name": "Security Activities",
            "tab_color": "C00000",
            "query": """
                SELECT sa.id, sa.name, sa.description, sa.frequency,
                       sa.owner_type, sa.owner_id, sa.created_at
                FROM security_activity sa
                ORDER BY sa.name
            """,
            "headers": ["ID", "Name", "Description", "Frequency",
                        "Owner Type", "Owner ID", "Created"],
            "status_cols": [],
            "text_cols": [2],
        },
        {
            "name": "Activity Executions",
            "tab_color": "C00000",
            "query": """
                SELECT ae.id, sa.name as activity,
                       u.name as executor, ae.execution_date,
                       ae.status, ae.outcome_notes, ae.created_at
                FROM activity_execution ae
                LEFT JOIN security_activity sa ON ae.activity_id = sa.id
                LEFT JOIN "user" u ON ae.executor_id = u.id
                ORDER BY ae.execution_date DESC
            """,
            "headers": ["ID", "Activity", "Executor", "Execution Date",
                        "Status", "Outcome Notes", "Created"],
            "status_cols": [4],
            "text_cols": [5],
        },

        # ── Security Assessments & Inventories ───────────────────────────
        {
            "name": "Security Assessments",
            "tab_color": "C00000",
            "query": """
                SELECT sa.id, sa.assessment_date, sa.status, sa.notes,
                       s.name as supplier
                FROM security_assessment sa
                LEFT JOIN supplier s ON sa.supplier_id = s.id
                ORDER BY sa.assessment_date DESC
            """,
            "headers": ["ID", "Assessment Date", "Status", "Notes", "Supplier"],
            "status_cols": [2],
            "text_cols": [3],
        },
        {
            "name": "Asset Inventories",
            "tab_color": "548235",
            "query": """
                SELECT ai.id, ai.name, ai.description,
                       u.name as conducted_by,
                       ai.is_completed, ai.created_at
                FROM asset_inventory ai
                LEFT JOIN "user" u ON ai.conducted_by_user_id = u.id
                ORDER BY ai.created_at DESC
            """,
            "headers": ["ID", "Name", "Description", "Conducted By",
                        "Completed", "Created"],
            "status_cols": [],
            "text_cols": [2],
        },

        # ── CRM / Requirements ───────────────────────────────────────────
        {
            "name": "Requirements",
            "tab_color": "BF8F00",
            "query": """
                SELECT r.id, r.name, r.requirement_type, r.priority,
                       r.status, r.description, r.estimated_budget,
                       r.currency, r.needed_by,
                       r.contact_name, r.email, r.phone,
                       u.name as created_by,
                       CASE WHEN r.is_archived THEN 'Archived' ELSE 'Active' END as archive_status,
                       r.created_at
                FROM lead r
                LEFT JOIN "user" u ON r.created_by_id = u.id
                ORDER BY r.created_at DESC
            """,
            "headers": ["ID", "Name", "Type", "Priority", "Status",
                        "Description", "Estimated Budget", "Currency",
                        "Needed By", "Contact Name", "Email", "Phone",
                        "Created By", "Archive Status", "Created"],
            "status_cols": [4, 13],
            "text_cols": [5],
        },
        {
            "name": "Opportunities",
            "tab_color": "BF8F00",
            "query": """
                SELECT o.id, o.name, o.status, o.potential_value, o.currency,
                       o.estimated_close_date, o.notes,
                       s.name as supplier, c.name as primary_contact,
                       o.created_at
                FROM opportunity o
                LEFT JOIN supplier s ON o.supplier_id = s.id
                LEFT JOIN contact c ON o.primary_contact_id = c.id
                ORDER BY o.created_at DESC
            """,
            "headers": ["ID", "Name", "Status", "Potential Value", "Currency",
                        "Est. Close Date", "Notes", "Supplier",
                        "Primary Contact", "Created"],
            "status_cols": [2],
            "text_cols": [6],
        },

        # ── Communications ───────────────────────────────────────────────
        {
            "name": "Email Templates",
            "tab_color": "00B0F0",
            "query": """
                SELECT id, name, subject, category,
                       is_active, is_system, created_at, updated_at
                FROM email_template
                ORDER BY category, name
            """,
            "headers": ["ID", "Name", "Subject", "Category",
                        "Active", "System", "Created", "Updated"],
            "status_cols": [],
            "text_cols": [],
        },
        {
            "name": "Campaigns",
            "tab_color": "00B0F0",
            "query": """
                SELECT ca.id, ca.title, ca.subject, ca.status,
                       ca.scheduled_at, ca.processed_at,
                       u.name as created_by,
                       ca.send_to_all, ca.created_at
                FROM campaign ca
                LEFT JOIN "user" u ON ca.created_by_id = u.id
                ORDER BY ca.created_at DESC
            """,
            "headers": ["ID", "Title", "Subject", "Status",
                        "Scheduled At", "Processed At", "Created By",
                        "Send To All", "Created"],
            "status_cols": [3],
            "text_cols": [],
        },

        # ── Risk Assessments ─────────────────────────────────────────────
        {
            "name": "Risk Assessments",
            "tab_color": "C00000",
            "query": """
                SELECT id, name, status,
                       total_residual_risk, locked_at, created_at
                FROM risk_assessment
                ORDER BY created_at DESC
            """,
            "headers": ["ID", "Name", "Status", "Total Residual Risk",
                        "Locked At", "Created"],
            "status_cols": [2],
            "text_cols": [],
        },

        # ── Organization Settings ────────────────────────────────────────
        {
            "name": "Org Settings",
            "tab_color": "2F5496",
            "query": """
                SELECT id, legal_name, tax_id, primary_domain,
                       logo_filename, email_domains, updated_at
                FROM organization_settings
            """,
            "headers": ["ID", "Legal Name", "Tax ID", "Primary Domain",
                        "Logo", "Email Domains", "Updated"],
            "status_cols": [],
            "text_cols": [],
        },
        {
            "name": "Cost Centers",
            "tab_color": "BF8F00",
            "query": """
                SELECT id, code, name, description, created_at
                FROM cost_center
                ORDER BY code
            """,
            "headers": ["ID", "Code", "Name", "Description", "Created"],
            "status_cols": [],
            "text_cols": [3],
        },
        {
            "name": "Onboarding Packs",
            "tab_color": "00B050",
            "query": """
                SELECT id, name, description, is_active
                FROM onboarding_pack
                ORDER BY name
            """,
            "headers": ["ID", "Name", "Description", "Active"],
            "status_cols": [],
            "text_cols": [2],
        },
    ]


# ─── DB Connection ───────────────────────────────────────────────────────────

def build_database_url(args):
    """Build the connection URL from the given arguments."""
    if args.database_url:
        return args.database_url

    db_url = os.environ.get("DATABASE_URL")
    if db_url:
        return db_url

    # Build from individual parameters
    return (
        f"postgresql://{args.user}:{args.password}"
        f"@{args.host}:{args.port}/{args.db}"
    )


def connect_db(database_url):
    """Create a SQLAlchemy connection to the database engine."""
    print(f"Connecting to: {database_url.split('@')[-1] if '@' in database_url else database_url}")
    try:
        engine = create_engine(database_url)
        # Test connection
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print("  Connection established.\n")
        return engine
    except Exception as e:
        print(f"Connection ERROR: {e}")
        sys.exit(1)


# ─── Excel Generation ────────────────────────────────────────────────────────

def write_sheet(wb, config, engine):
    """
    Write a workbook sheet with the query data.
    Returns the number of data rows written.
    """
    name = config["name"]
    headers = config["headers"]
    status_cols = config.get("status_cols", [])
    text_cols = config.get("text_cols", [])
    tab_color = config.get("tab_color", "2F5496")

    # Use the active sheet (each file has a single sheet)
    ws = wb.active
    ws.title = name
    ws.sheet_properties.tabColor = tab_color

    num_cols = len(headers)

    # ── Title ──
    if num_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=f"OpsDeck BCDR — {name}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = TITLE_ALIGNMENT
    ws.row_dimensions[1].height = 30

    # ── Export info ──
    if num_cols > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    info_cell = ws.cell(
        row=2, column=1,
        value=f"Exported on {datetime.now().strftime('%d/%m/%Y at %H:%M')}"
    )
    info_cell.font = INFO_FONT
    info_cell.alignment = INFO_ALIGNMENT

    # ── Headers ──
    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=header)
    apply_header_style(ws, header_row, num_cols)
    ws.row_dimensions[header_row].height = 25

    # ── Data ──
    row_count = 0
    try:
        with engine.connect() as conn:
            result = conn.execute(text(config["query"]))
            rows = result.fetchall()

        if not rows:
            r = header_row + 1
            if num_cols > 1:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
            empty_cell = ws.cell(row=r, column=1, value="No data")
            empty_cell.font = EMPTY_FONT
            empty_cell.alignment = EMPTY_ALIGNMENT
        else:
            for i, row_data in enumerate(rows):
                r = header_row + 1 + i
                for col_idx in range(num_cols):
                    val = fmt_val(row_data[col_idx]) if col_idx < len(row_data) else ""
                    cell = ws.cell(row=r, column=col_idx + 1, value=val)
                    center = col_idx not in text_cols
                    apply_data_style(cell, center=center)

                    # Apply status color
                    if col_idx in status_cols:
                        apply_status_color(cell, val)

                row_count += 1

    except Exception as e:
        r = header_row + 1
        if num_cols > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        err_cell = ws.cell(row=r, column=1, value=f"Error: {e}")
        err_cell.font = Font(name="Calibri", italic=True, color="FF0000", size=10)
        err_cell.alignment = EMPTY_ALIGNMENT
        print(f"    Query ERROR: {e}")

    auto_width(ws)
    return row_count


def create_index_file(output_dir, sheet_configs, row_counts):
    """Create an index Excel file with the export summary."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    ws.sheet_properties.tabColor = "000000"

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "OpsDeck BCDR Export — Index"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    title_cell.alignment = TITLE_ALIGNMENT
    ws.row_dimensions[1].height = 35

    # Info
    ws.merge_cells("A2:E2")
    info_cell = ws["A2"]
    info_cell.value = f"Generated on {datetime.now().strftime('%d/%m/%Y at %H:%M')}"
    info_cell.font = INFO_FONT
    info_cell.alignment = INFO_ALIGNMENT

    ws.merge_cells("A3:E3")
    total_cell = ws["A3"]
    total_records = sum(row_counts.values())
    total_cell.value = f"Total: {len(sheet_configs)} files | {total_records:,} records"
    total_cell.font = Font(name="Calibri", bold=True, size=11, color="333333")
    total_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["#", "File", "Records", "Category", "Status"]
    row = 5
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    apply_header_style(ws, row, len(headers))

    # Categories by color
    TAB_CATEGORIES = {
        "2F5496": "Organization",
        "548235": "Assets & Inventory",
        "BF8F00": "Procurement & Vendors",
        "7030A0": "Services & Credentials",
        "C00000": "Security & Risks",
        "FF6600": "Incidents & Changes",
        "4472C4": "Compliance & Audits",
        "00B050": "BCDR & HR",
        "00B0F0": "Documentation & Comms",
    }

    for i, config in enumerate(sheet_configs):
        r = row + 1 + i
        name = config["name"]
        count = row_counts.get(name, 0)
        safe_name = name.replace(" ", "_").replace("/", "_")

        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=f"{safe_name}.xlsx")
        ws.cell(row=r, column=3, value=count)
        category = TAB_CATEGORIES.get(config.get("tab_color", ""), "Other")
        ws.cell(row=r, column=4, value=category)
        status = "Has data" if count > 0 else "Empty"
        ws.cell(row=r, column=5, value=status)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            apply_data_style(cell, center=(col != 2))
            if i % 2 == 0:
                cell.fill = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )
        if count > 0:
            apply_status_color(ws.cell(row=r, column=5), "active")
        else:
            apply_status_color(ws.cell(row=r, column=5), "pending")

    auto_width(ws)

    filepath = output_dir / "_Index.xlsx"
    wb.save(filepath)
    return filepath


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="BCDR Export - Extract all OpsDeck data to Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/bcdr_export.py --database-url postgresql://opsdeck:opsdeck@localhost:5432/opsdeck
  python scripts/bcdr_export.py --host db.prod.internal --db opsdeck_prod
  python scripts/bcdr_export.py --output /backups/opsdeck_$(date +%%Y%%m%%d)
        """,
    )
    parser.add_argument(
        "--database-url", default=None,
        help="Full connection URL (default: $DATABASE_URL)"
    )
    parser.add_argument(
        "--host", default=os.getenv("POSTGRES_HOST", "localhost"),
        help="PostgreSQL host (default: $POSTGRES_HOST or localhost)"
    )
    parser.add_argument(
        "--port", default=os.getenv("POSTGRES_PORT", "5432"),
        help="PostgreSQL port (default: 5432)"
    )
    parser.add_argument(
        "--db", default=os.getenv("POSTGRES_DB", "opsdeck"),
        help="Database name (default: $POSTGRES_DB or opsdeck)"
    )
    parser.add_argument(
        "--user", default=os.getenv("POSTGRES_USER", "opsdeck"),
        help="DB user (default: $POSTGRES_USER or opsdeck)"
    )
    parser.add_argument(
        "--password", default=os.getenv("POSTGRES_PASSWORD", "opsdeck"),
        help="DB password (default: $POSTGRES_PASSWORD or opsdeck)"
    )
    parser.add_argument(
        "--output", "-o", default=None,
        help="Output directory (default: bcdr_opsdeck_YYYYMMDD_HHMM/)"
    )

    args = parser.parse_args()

    # Output directory
    if args.output:
        output_dir = Path(args.output)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_dir = Path(f"bcdr_opsdeck_{timestamp}")

    output_dir.mkdir(parents=True, exist_ok=True)

    # Connection
    database_url = build_database_url(args)
    engine = connect_db(database_url)

    # Sheet configuration
    sheet_configs = get_sheet_configs()

    print(f"Exporting {len(sheet_configs)} files to {output_dir.resolve()}/")
    print("=" * 60)

    row_counts = {}

    for i, config in enumerate(sheet_configs):
        name = config["name"]
        safe_name = name.replace(" ", "_").replace("/", "_")

        wb = Workbook()
        count = write_sheet(wb, config, engine)
        row_counts[name] = count

        filepath = output_dir / f"{safe_name}.xlsx"
        wb.save(filepath)

        status = f"{count:>6,} records" if count > 0 else "      (empty)"
        print(f"  [{i+1:2d}/{len(sheet_configs)}] {safe_name + '.xlsx':<35s} {status}")

    # Create index file
    index_path = create_index_file(output_dir, sheet_configs, row_counts)

    print("\n" + "=" * 60)

    total = sum(row_counts.values())
    non_empty = sum(1 for c in row_counts.values() if c > 0)
    total_size = sum(f.stat().st_size for f in output_dir.glob("*.xlsx"))
    print(f"\nExport completed:")
    print(f"  Directory: {output_dir.resolve()}")
    print(f"  Files:     {len(sheet_configs) + 1} ({non_empty} with data + index)")
    print(f"  Records:   {total:,}")
    print(f"  Size:      {total_size / 1024:.1f} KB")

    engine.dispose()


if __name__ == "__main__":
    main()
